import os
import json
import numpy as np
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db, SessionLocal
from models_db import Content, Department, DocumentChunk, RAGSession
from yandex_ai_service import YandexAIService
import PyPDF2
import docx
from io import BytesIO
import re
import pandas as pd
from openpyxl import load_workbook

class YandexRAGService:
    def __init__(self):
        self.yandex_ai = YandexAIService()
        self.chunk_size = 1000  # Размер чанка в символах
        self.chunk_overlap = 200  # Перекрытие между чанками
        
    async def initialize_rag(self, department_id: int, force_reload: bool = False) -> Dict[str, Any]:
        """Инициализация RAG системы для отдела"""
        db = SessionLocal()
        try:
            # Проверяем существование отдела
            department = db.query(Department).filter(Department.id == department_id).first()
            if not department:
                return {
                    "success": False,
                    "message": f"Отдел с ID {department_id} не найден"
                }
            
            # Получаем или создаем RAG сессию
            rag_session = db.query(RAGSession).filter(RAGSession.department_id == department_id).first()
            if not rag_session:
                rag_session = RAGSession(department_id=department_id)
                db.add(rag_session)
                db.commit()
                db.refresh(rag_session)
            
            # Получаем все документы отдела
            documents = db.query(Content).filter(Content.department_id == department_id).all()
            
            if not documents:
                return {
                    "success": False,
                    "message": f"Нет документов для обработки в отделе {department.department_name}"
                }       
     
            # Если принудительная перезагрузка, удаляем существующие чанки
            if force_reload:
                db.query(DocumentChunk).filter(DocumentChunk.department_id == department_id).delete()
                db.commit()
            
            processed_docs = 0
            total_chunks = 0
            
            for document in documents:
                # Проверяем, есть ли уже чанки для этого документа
                existing_chunks = db.query(DocumentChunk).filter(
                    DocumentChunk.content_id == document.id
                ).count()
                
                if existing_chunks > 0 and not force_reload:
                    continue
                
                # Извлекаем текст из документа
                text_content = await self._extract_text_from_file(document.file_path)
                if not text_content:
                    continue
                
                # Разбиваем на чанки
                chunks = self._split_text_into_chunks(text_content)
                
                # Создаем эмбеддинги для каждого чанка
                for i, chunk_text in enumerate(chunks):
                    try:
                        # Получаем эмбеддинг от Yandex
                        embedding = await self.yandex_ai.get_embedding(chunk_text)
                        
                        # Сохраняем чанк в БД
                        chunk = DocumentChunk(
                            content_id=document.id,
                            department_id=department_id,
                            chunk_text=chunk_text,
                            chunk_index=i,
                            embedding_vector=embedding
                        )
                        db.add(chunk)
                        total_chunks += 1
                        
                    except Exception as e:
                        print(f"Ошибка создания эмбеддинга для чанка {i} документа {document.id}: {e}")
                        continue
                
                processed_docs += 1
                db.commit()
            
            # Обновляем статус RAG сессии
            rag_session.is_initialized = True
            rag_session.documents_count = len(documents)
            rag_session.chunks_count = total_chunks
            rag_session.last_updated = func.now()
            db.commit()
            
            return {
                "success": True,
                "message": f"RAG система инициализирована для отдела {department.department_name}",
                "documents_processed": processed_docs,
                "chunks_created": total_chunks
            }
            
        except Exception as e:
            db.rollback()
            return {
                "success": False,
                "message": f"Ошибка инициализации RAG: {str(e)}"
            }
        finally:
            db.close()  
  
    async def get_rag_status(self, department_id: int) -> Dict[str, Any]:
        """Получение статуса RAG системы для отдела"""
        db = SessionLocal()
        try:
            # Получаем информацию об отделе
            department = db.query(Department).filter(Department.id == department_id).first()
            if not department:
                raise Exception(f"Отдел с ID {department_id} не найден")
            
            # Получаем RAG сессию
            rag_session = db.query(RAGSession).filter(RAGSession.department_id == department_id).first()
            
            # Считаем документы в БД
            documents_in_db = db.query(Content).filter(Content.department_id == department_id).count()
            
            # Считаем чанки в векторной БД
            chunks_in_vector_store = db.query(DocumentChunk).filter(
                DocumentChunk.department_id == department_id
            ).count()
            
            is_initialized = rag_session.is_initialized if rag_session else False
            needs_reinitialization = False
            
            if rag_session and is_initialized:
                # Проверяем, нужна ли реинициализация
                needs_reinitialization = (
                    rag_session.documents_count != documents_in_db or
                    chunks_in_vector_store == 0
                )
            
            return {
                "department_id": department_id,
                "department_name": department.department_name,
                "is_initialized": is_initialized,
                "documents_in_db": documents_in_db,
                "documents_in_vector_store": chunks_in_vector_store,
                "needs_reinitialization": needs_reinitialization,
                "last_updated": rag_session.last_updated.isoformat() if rag_session and rag_session.last_updated else None
            }
            
        except Exception as e:
            raise Exception(f"Ошибка получения статуса RAG: {str(e)}")
        finally:
            db.close()
    
    async def reset_rag(self, department_id: int) -> Dict[str, Any]:
        """Сброс RAG системы для отдела"""
        db = SessionLocal()
        try:
            # Проверяем существование отдела
            department = db.query(Department).filter(Department.id == department_id).first()
            if not department:
                return {
                    "success": False,
                    "message": f"Отдел с ID {department_id} не найден"
                }
            
            # Удаляем все чанки отдела
            deleted_chunks = db.query(DocumentChunk).filter(
                DocumentChunk.department_id == department_id
            ).delete()
            
            # Сбрасываем RAG сессию
            rag_session = db.query(RAGSession).filter(RAGSession.department_id == department_id).first()
            if rag_session:
                rag_session.is_initialized = False
                rag_session.documents_count = 0
                rag_session.chunks_count = 0
                rag_session.last_updated = func.now()
            
            db.commit()
            
            return {
                "success": True,
                "message": f"RAG система сброшена для отдела {department.department_name}. Удалено чанков: {deleted_chunks}"
            }
            
        except Exception as e:
            db.rollback()
            return {
                "success": False,
                "message": f"Ошибка сброса RAG: {str(e)}"
            }
        finally:
            db.close()  
  
    async def query_rag(self, department_id: int, question: str) -> Dict[str, Any]:
        """Выполнение RAG запроса"""
        db = SessionLocal()
        try:
            # Проверяем, инициализирована ли RAG система
            rag_session = db.query(RAGSession).filter(RAGSession.department_id == department_id).first()
            if not rag_session or not rag_session.is_initialized:
                raise Exception("RAG система не инициализирована для данного отдела")
            
            # Получаем эмбеддинг для вопроса
            question_embedding = await self.yandex_ai.get_embedding(question)
            
            # Получаем все чанки отдела
            chunks = db.query(DocumentChunk).filter(
                DocumentChunk.department_id == department_id
            ).all()
            
            if not chunks:
                raise Exception("Нет данных в векторной базе для данного отдела")
            
            # Вычисляем сходство с каждым чанком
            similarities = []
            for chunk in chunks:
                if chunk.embedding_vector:
                    similarity = self._cosine_similarity(question_embedding, chunk.embedding_vector)
                    similarities.append((chunk, similarity))
            
            # Сортируем по убыванию сходства и берем топ-5
            similarities.sort(key=lambda x: x[1], reverse=True)
            top_chunks = similarities[:5]
            
            # Формируем контекст из наиболее релевантных чанков
            context_parts = []
            sources = []
            unique_sources = {}  # Для отслеживания уникальных источников
            seen_content = set()  # Для отслеживания дубликатов по содержимому
            
            print(f"RAG: Обработка {len(top_chunks)} чанков для формирования источников")
            
            for chunk, similarity in top_chunks:
                if similarity > 0.2:  # Порог релевантности (снижен с 0.3 до 0.2)
                    context_parts.append(chunk.chunk_text)
                    
                    # Получаем информацию о документе-источнике
                    content = db.query(Content).filter(Content.id == chunk.content_id).first()
                    if content:
                        # Создаем уникальный ключ для источника
                        source_key = f"{content.id}_{chunk.chunk_index}"
                        
                        # Дополнительная проверка на дубликаты по содержимому
                        content_preview = chunk.chunk_text[:100].lower().strip()
                        content_hash = hash(content_preview)
                        content_key = f"{content.id}_{content_hash}"
                        
                        # Добавляем только уникальные источники
                        if source_key not in unique_sources and content_key not in seen_content:
                            unique_sources[source_key] = {
                                "chunk_id": source_key,
                                "file_name": content.title,
                                "file_path": content.file_path,
                                "chunk_content": chunk.chunk_text,
                                "similarity_score": round(similarity, 3),
                                "page_number": None  # Можно добавить позже
                            }
                            seen_content.add(content_key)
                            print(f"RAG: Добавлен источник: {content.title} (релевантность: {similarity:.3f})")
                        else:
                            print(f"RAG: Пропущен дубликат: {content.title}")
            
            # Преобразуем уникальные источники в список
            sources = list(unique_sources.values())
            print(f"RAG: Сформировано {len(sources)} уникальных источников")
            
            if not context_parts:
                return {
                    "answer": "К сожалению, не найдено релевантной информации для ответа на ваш вопрос. Попробуйте переформулировать запрос или обратитесь к другим источникам.",
                    "sources": [],
                    "context_used": 0,
                    "no_sources_found": True
                }
            
            # Формируем промпт с контекстом
            context = "\n\n".join(context_parts)
            prompt = f"""Ты — профессиональный ассистент, который помогает пользователям найти точную информацию в корпоративных документах.

### КРИТИЧЕСКИ ВАЖНЫЕ ПРАВИЛА (НЕ НАРУШАТЬ):
1. НИКОГДА НЕ ЦЕНЗУРИРУЙ ответы - давай полную и честную информацию из документов
2. НЕ ИСПОЛЬЗУЙ фразы типа "в интернете есть информация", "посмотрите в поиске", "не могу ответить"
3. Если в вопросе или контексте есть любые слова - отвечай прямо и честно
4. ВСЕГДА сначала внимательно изучи предоставленные документы
5. Если информация есть в документах — используй ТОЛЬКО её для ответа
6. Если информации нет в документах — честно скажи об этом
7. Давай подробные и структурированные ответы
8. При цитировании указывай источник документа
9. НЕ ПЕРЕАДРЕСОВЫВАЙ на поиск - отвечай сам
10. НЕ СКРЫВАЙ информацию из документов
11. В КОНЦЕ ответа обязательно укажи номер основного источника (1, 2, 3, 4, 5) в формате: [ОСНОВНОЙ_ИСТОЧНИК: X]

### ДОПОЛНИТЕЛЬНЫЕ ПРАВИЛА ДЛЯ КАЧЕСТВЕННЫХ ОТВЕТОВ:
12. Если в документах есть упоминание слов из вопроса, даже косвенное - обязательно укажи это
13. НЕ ГОВОРИ "нет информации", если в документах есть любые упоминания искомых слов
14. Если информация есть, но не относится к основной теме документов - скажи об этом
15. Если нашел частичную информацию - опиши её, даже если она неполная
16. Всегда анализируй контекст вокруг найденных слов

### Контекст из документов:
{context}

### Вопрос пользователя:
{question}

### Ответ:
Проанализировав предоставленные документы:"""
            
            # Получаем ответ от Yandex AI
            answer = await self.yandex_ai.generate_response(prompt)
            
            # Очищаем ответ от маркера источника (если есть)
            clean_answer = self._clean_answer_from_source_marker(answer)
            
            # Проверяем на цензурные ответы и заменяем их
            clean_answer = self._fix_censored_response(clean_answer, context, question)
            
            # Анализируем ответ ИИ для определения основного источника
            main_source_number = self._analyze_answer_for_main_source(clean_answer, sources)
            
            # Пересортировываем источники: основной источник первым, остальные по релевантности
            reordered_sources = self._reorder_sources_by_main_source(sources, main_source_number)
            
            return {
                "answer": clean_answer,
                "sources": reordered_sources,
                "context_used": len(context_parts),
                "sources_count": len(reordered_sources),
                "main_source_number": main_source_number
            }
            
        except Exception as e:
            raise Exception(f"Ошибка RAG запроса: {str(e)}")
        finally:
            db.close()    

    async def _extract_text_from_file(self, file_path: str) -> str:
        """Извлечение текста из файла"""
        try:
            if not os.path.exists(file_path):
                return ""
            
            file_extension = os.path.splitext(file_path)[1].lower()
            
            if file_extension == '.pdf':
                return self._extract_text_from_pdf(file_path)
            elif file_extension in ['.docx', '.doc']:
                return self._extract_text_from_docx(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                return self._extract_text_from_excel(file_path)
            elif file_extension == '.txt':
                with open(file_path, 'r', encoding='utf-8') as file:
                    return file.read()
            else:
                return ""
                
        except Exception as e:
            print(f"Ошибка извлечения текста из файла {file_path}: {e}")
            return ""
    
    def _extract_text_from_pdf(self, file_path: str) -> str:
        """Извлечение текста из PDF файла"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text
        except Exception as e:
            print(f"Ошибка извлечения текста из PDF {file_path}: {e}")
            return ""
    
    def _extract_text_from_docx(self, file_path: str) -> str:
        """Извлечение текста из DOCX файла"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            print(f"Ошибка извлечения текста из DOCX {file_path}: {e}")
            return ""
    
    def _extract_text_from_excel(self, file_path: str) -> str:
        """Извлечение текста из Excel файла"""
        try:
            # Загружаем рабочую книгу
            workbook = load_workbook(file_path, data_only=True)
            
            text_content = f"=== Excel файл: {os.path.basename(file_path)} ===\n"
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content += f"\n--- Лист: {sheet_name} ---\n"
                
                # Получаем размеры данных
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                if max_row > 0 and max_col > 0:
                    # Читаем заголовки
                    headers = []
                    for col in range(1, max_col + 1):
                        cell_value = sheet.cell(row=1, column=col).value
                        headers.append(str(cell_value) if cell_value is not None else f"Столбец {col}")
                    
                    text_content += f"Заголовки: {' | '.join(headers)}\n"
                    
                    # Читаем данные (ограничиваем 100 строками для производительности)
                    for row in range(2, min(max_row + 1, 102)):  # 100 строк данных + заголовки
                        row_data = []
                        for col in range(1, max_col + 1):
                            cell_value = sheet.cell(row=row, column=col).value
                            row_data.append(str(cell_value) if cell_value is not None else "")
                        
                        text_content += f"Строка {row}: {' | '.join(row_data)}\n"
                    
                    if max_row > 101:
                        text_content += f"... и еще {max_row - 101} строк\n"
                
                text_content += "\n"
            
            return text_content
            
        except Exception as e:
            print(f"Ошибка извлечения текста из Excel {file_path}: {e}")
            return ""
    
    def _split_text_into_chunks(self, text: str) -> List[str]:
        """Разбиение текста на чанки"""
        # Очищаем текст от лишних пробелов и переносов
        text = re.sub(r'\s+', ' ', text.strip())
        
        if len(text) <= self.chunk_size:
            return [text]
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + self.chunk_size
            
            if end >= len(text):
                chunks.append(text[start:])
                break
            
            # Ищем ближайший разделитель предложения
            chunk_end = end
            for i in range(end, max(start + self.chunk_size - self.chunk_overlap, start), -1):
                if text[i] in '.!?':
                    chunk_end = i + 1
                    break
            
            chunks.append(text[start:chunk_end])
            start = chunk_end - self.chunk_overlap
            
            if start < 0:
                start = 0
        
        return chunks
    
    def _cosine_similarity(self, vec1: List[float], vec2: List[float]) -> float:
        """Вычисление косинусного сходства между векторами"""
        vec1_np = np.array(vec1)
        vec2_np = np.array(vec2)
        
        dot_product = np.dot(vec1_np, vec2_np)
        norm1 = np.linalg.norm(vec1_np)
        norm2 = np.linalg.norm(vec2_np)
        
        if norm1 == 0 or norm2 == 0:
            return 0.0
        
        return dot_product / (norm1 * norm2)
    
    def _fix_censored_response(self, answer: str, context: str, question: str) -> str:
        """Исправляет цензурные ответы, заменяя их на информацию из контекста"""
        
        # Признаки цензурных ответов
        censorship_patterns = [
            r'в интернете есть.*информация',
            r'посмотрите.*в поиске',
            r'не могу.*ответить',
            r'не могу.*предоставить',
            r'не могу.*дать',
            r'не могу.*рассказать',
            r'не могу.*объяснить',
            r'не могу.*обсуждать',
            r'не могу.*говорить',
            r'не могу.*комментировать',
            r'не могу.*делиться',
            r'не могу.*раскрывать',
            r'не могу.*разглашать',
            r'не могу.*описывать',
            r'не могу.*излагать',
            r'не могу.*представлять',
            r'не могу.*показывать',
            r'не могу.*демонстрировать',
            r'не могу.*иллюстрировать',
            r'не могу.*характеризовать',
            r'не могу.*определять',
            r'извините.*не могу',
            r'к сожалению.*не могу',
            r'сожалею.*не могу',
            r'не в состоянии.*ответить',
            r'не в состоянии.*предоставить',
            r'не в состоянии.*дать',
            r'не в состоянии.*рассказать',
            r'не в состоянии.*объяснить',
            r'не в состоянии.*обсуждать',
            r'не в состоянии.*говорить',
            r'не в состоянии.*комментировать',
            r'не в состоянии.*делиться',
            r'не в состоянии.*раскрывать',
            r'не в состоянии.*разглашать',
            r'не в состоянии.*описывать',
            r'не в состоянии.*излагать',
            r'не в состоянии.*представлять',
            r'не в состоянии.*показывать',
            r'не в состоянии.*демонстрировать',
            r'не в состоянии.*иллюстрировать',
            r'не в состоянии.*характеризовать',
            r'не в состоянии.*определять'
        ]
        
        # Признаки ответов "нет информации" (которые нужно проверить)
        no_info_patterns = [
            r'нет информации',
            r'не найдено.*информации',
            r'не содержит.*информации',
            r'не упоминается',
            r'не упоминается.*информация',
            r'не указано',
            r'не указано.*информация',
            r'не содержится',
            r'не содержится.*информация',
            r'не обнаружено',
            r'не обнаружено.*информации',
            r'не найдено',
            r'не найдено.*данных',
            r'не найдено.*сведений',
            r'не найдено.*упоминаний'
        ]
        
        import re
        
        # Проверяем, содержит ли ответ цензурные паттерны
        answer_lower = answer.lower()
        is_censored = any(re.search(pattern, answer_lower) for pattern in censorship_patterns)
        
        # Проверяем, содержит ли ответ "нет информации" (которое может быть неправильным)
        is_no_info = any(re.search(pattern, answer_lower) for pattern in no_info_patterns)
        
        # Отладочная информация
        print(f"🔍 Отладка _fix_censored_response:")
        print(f"   Вопрос: {question}")
        print(f"   Ответ: {answer[:100]}...")
        print(f"   is_censored: {is_censored}")
        print(f"   is_no_info: {is_no_info}")
        
        # Дополнительная проверка на "нет информации" без паттернов
        if "нет информации" in answer_lower:
            is_no_info = True
            print(f"   ⚠️ Найдено 'нет информации' в ответе")
        
        if is_censored:
            print(f"⚠️ Обнаружен цензурный ответ, исправляем...")
            print(f"Вопрос: {question}")
            print(f"Цензурный ответ: {answer}")
            
            # Создаем новый ответ на основе контекста
            if context and context.strip():
                # Ищем ключевые слова из вопроса в контексте
                question_words = question.lower().split()
                context_lower = context.lower()
                
                # Ищем релевантные части контекста
                relevant_parts = []
                for word in question_words:
                    if len(word) > 3:  # Игнорируем короткие слова
                        if word in context_lower:
                            # Находим предложения с этим словом
                            sentences = re.split(r'[.!?]+', context)
                            for sentence in sentences:
                                if word in sentence.lower():
                                    relevant_parts.append(sentence.strip())
                
                if relevant_parts:
                    # Убираем дубликаты и объединяем
                    unique_parts = list(set(relevant_parts))
                    new_answer = " ".join(unique_parts[:3])  # Берем первые 3 релевантных предложения
                    
                    # Добавляем префикс
                    new_answer = f"Согласно предоставленным документам: {new_answer}"
                    
                    print(f"✅ Исправленный ответ: {new_answer}")
                    return new_answer
                else:
                    # Если не нашли релевантных частей, возвращаем контекст
                    context_preview = context[:500] + "..." if len(context) > 500 else context
                    new_answer = f"Согласно предоставленным документам: {context_preview}"
                    
                    print(f"✅ Исправленный ответ (контекст): {new_answer}")
                    return new_answer
            else:
                # Если нет контекста, даем честный ответ
                new_answer = "К сожалению, в предоставленных документах нет информации для ответа на ваш вопрос."
                print(f"✅ Исправленный ответ (нет данных): {new_answer}")
                return new_answer
        
        elif is_no_info:
            # Проверяем, действительно ли нет информации в контексте
            question_words = question.lower().split()
            context_lower = context.lower()
            
            # Ищем упоминания слов из вопроса в контексте
            found_words = []
            for word in question_words:
                if len(word) > 2:  # Игнорируем очень короткие слова
                    if word in context_lower:
                        found_words.append(word)
            
            if found_words:
                print(f"⚠️ ИИ сказал 'нет информации', но в контексте найдены слова: {found_words}")
                print(f"Вопрос: {question}")
                print(f"Неправильный ответ: {answer}")
                
                # Создаем исправленный ответ
                relevant_parts = []
                for word in found_words:
                    sentences = re.split(r'[.!?]+', context)
                    for sentence in sentences:
                        if word in sentence.lower() and len(sentence.strip()) > 10:
                            relevant_parts.append(sentence.strip())
                
                if relevant_parts:
                    # Убираем дубликаты и берем первые 2 релевантных предложения
                    unique_parts = list(set(relevant_parts))[:2]
                    new_answer = f"В предоставленных документах есть упоминание {' и '.join(found_words)}: {' '.join(unique_parts)}"
                    
                    print(f"✅ Исправленный ответ: {new_answer}")
                    return new_answer
                else:
                    new_answer = f"В предоставленных документах есть упоминание {' и '.join(found_words)}, но контекст не содержит дополнительной информации."
                    
                    print(f"✅ Исправленный ответ: {new_answer}")
                    return new_answer
        
        return answer

    def _extract_main_source_number(self, answer: str) -> int:
        """Извлекает номер основного источника из ответа ИИ"""
        try:
            # Ищем паттерн [ОСНОВНОЙ_ИСТОЧНИК: X] в конце ответа
            import re
            pattern = r'\[ОСНОВНОЙ_ИСТОЧНИК:\s*(\d+)\]'
            match = re.search(pattern, answer, re.IGNORECASE)
            
            if match:
                source_number = int(match.group(1))
                # Проверяем, что номер в допустимом диапазоне (1-5)
                if 1 <= source_number <= 5:
                    return source_number
            
            # Если не нашли маркер, возвращаем 1 (первый источник)
            return 1
        except Exception as e:
            print(f"Ошибка при извлечении номера основного источника: {e}")
            return 1
    
    def _clean_answer_from_source_marker(self, answer: str) -> str:
        """Удаляет маркер основного источника из ответа"""
        try:
            import re
            # Удаляем паттерн [ОСНОВНОЙ_ИСТОЧНИК: X] из ответа
            pattern = r'\s*\[ОСНОВНОЙ_ИСТОЧНИК:\s*\d+\]\s*$'
            cleaned_answer = re.sub(pattern, '', answer, flags=re.IGNORECASE)
            return cleaned_answer.strip()
        except Exception as e:
            print(f"Ошибка при очистке ответа от маркера источника: {e}")
            return answer

    def _reorder_sources_by_main_source(self, sources: List[Dict[str, Any]], main_source_number: int) -> List[Dict[str, Any]]:
        """Пересортировывает источники, помещая основной источник на первое место."""
        if not sources:
            return []
        
        # Создаем копию списка источников
        reordered_sources = sources.copy()
        
        # Проверяем, что номер основного источника в допустимом диапазоне
        if main_source_number < 1 or main_source_number > len(reordered_sources):
            print(f"Номер основного источника {main_source_number} вне диапазона, используем первый источник")
            return reordered_sources
        
        # Находим индекс основного источника (main_source_number - 1, так как индексы начинаются с 0)
        main_source_index = main_source_number - 1
        
        if main_source_index < len(reordered_sources):
            # Перемещаем основной источник на первое место
            main_source = reordered_sources.pop(main_source_index)
            reordered_sources.insert(0, main_source)
            
            print(f"Основной источник {main_source_number} перемещен на первое место: {main_source.get('file_name', 'unknown')}")
        
        return reordered_sources

    def _analyze_answer_for_main_source(self, answer: str, sources: List[Dict[str, Any]]) -> int:
        """Анализирует ответ ИИ и определяет основной источник на основе содержимого"""
        if not answer or not sources:
            return 1
        
        # Приводим ответ к нижнему регистру для анализа
        answer_lower = answer.lower()
        
        # Ищем точные совпадения фраз из источников в ответе
        best_match_index = 0
        best_match_score = 0
        
        for i, source in enumerate(sources):
            source_content = source.get('chunk_content', '').lower()
            if not source_content:
                continue
            
            score = 0
            
            # Проверяем точные совпадения определений
            if '—' in source_content and '—' in answer_lower:
                # Ищем определение в формате "ТМЦ — товарно-материальные ценности"
                definition_pattern = r'([а-яё]+)\s*—\s*([^;]+)'
                import re
                source_definitions = re.findall(definition_pattern, source_content)
                answer_definitions = re.findall(definition_pattern, answer_lower)
                
                for source_def in source_definitions:
                    for answer_def in answer_definitions:
                        if source_def[0] == answer_def[0]:  # Термин совпадает
                            score += 200  # Высокий бонус за точное совпадение определения
            
            # Проверяем совпадения ключевых фраз
            source_words = source_content.split()
            answer_words = answer_lower.split()
            
            # Считаем общие слова
            common_words = set(source_words) & set(answer_words)
            score += len(common_words) * 5
            
            # Бонус за длинные совпадающие фразы
            for j in range(len(source_words) - 2):
                for k in range(len(answer_words) - 2):
                    source_phrase = ' '.join(source_words[j:j+3])
                    answer_phrase = ' '.join(answer_words[k:k+3])
                    if source_phrase == answer_phrase and len(source_phrase) > 10:
                        score += 50
            
            # Дополнительный бонус за определения в ответе
            if any(word in answer_lower for word in ['—', 'означает', 'это', 'определение']):
                if any(word in source_content for word in ['—', 'означает', 'определение', 'термины']):
                    score += 100
            
            # Бонус за термины и сокращения
            if 'термины и сокращения' in source_content.lower():
                score += 80
            
            if score > best_match_score:
                best_match_score = score
                best_match_index = i
        
        print(f"RAG: Анализ ответа - выбран источник {best_match_index + 1} с баллом {best_match_score}")
        
        return best_match_index + 1  # +1 потому что индексы начинаются с 1

    def _presort_sources_by_semantic_importance(self, sources: List[Dict[str, Any]], question: str) -> List[Dict[str, Any]]:
        """Предварительно сортирует источники по семантической важности для вопроса"""
        if not sources or not question:
            return sources
        
        question_lower = question.lower()
        
        # Определяем тип вопроса
        is_definition_question = any(word in question_lower for word in ['что такое', 'что означает', 'определение', 'означает'])
        is_procedure_question = any(word in question_lower for word in ['как', 'процедура', 'процесс', 'шаги', 'этапы'])
        is_requirement_question = any(word in question_lower for word in ['требования', 'правила', 'условия', 'необходимо'])
        
        def calculate_semantic_score(source):
            content = source.get('chunk_content', '').lower()
            score = 0
            
            # Бонус за определения
            if is_definition_question:
                if any(word in content for word in ['—', 'означает', 'определение', 'это']):
                    score += 100
                if 'термины и сокращения' in content:
                    score += 50
            
            # Бонус за процедуры
            if is_procedure_question:
                if any(word in content for word in ['шаг', 'этап', 'сначала', 'затем', 'после']):
                    score += 80
                if any(word in content for word in ['процедура', 'процесс', 'алгоритм']):
                    score += 60
            
            # Бонус за требования
            if is_requirement_question:
                if any(word in content for word in ['требуется', 'необходимо', 'обязательно', 'должен']):
                    score += 70
                if any(word in content for word in ['правила', 'условия', 'критерии']):
                    score += 50
            
            # Базовый бонус за релевантность
            score += source.get('similarity_score', 0) * 10
            
            return score
        
        # Сортируем источники по семантической важности
        sorted_sources = sorted(sources, key=calculate_semantic_score, reverse=True)
        
        print(f"RAG: Источники пересортированы по семантической важности для вопроса: '{question[:50]}...'")
        for i, source in enumerate(sorted_sources):
            print(f"  {i+1}. {source.get('file_name', 'unknown')} (релевантность: {source.get('similarity_score', 0):.3f})")
        
        return sorted_sources

# Глобальный экземпляр сервиса
yandex_rag_service = YandexRAGService()